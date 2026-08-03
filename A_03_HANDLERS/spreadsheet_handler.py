import csv
import os
import tempfile
from pathlib import Path
from A_03_HANDLERS.base_handler import BaseHandler

class SpreadsheetHandler(BaseHandler):

    supported_extensions = [".csv", ".xlsx"]

    def create_csv(self, target, rows):
        target = Path(str(target))
        if not target.is_absolute() or target.suffix.lower() != ".csv":
            return {"success": False, "error": "INVALID_PATH", "text": "Требуется абсолютный Windows-путь к файлу .csv.", "metadata": {}}
        if target.exists():
            return {"success": False, "error": "CSV_TARGET_EXISTS", "text": f"Целевой файл уже существует: {target}", "metadata": {}}
        if not rows:
            return {"success": False, "error": "CREATE_FAILED", "text": "Не указаны данные для CSV.", "metadata": {}}
        if not rows[0] or any(len(row) != len(rows[0]) for row in rows):
            return {"success": False, "error": "CREATE_FAILED", "text": "Табличные данные имеют разное число столбцов.", "metadata": {}}

        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            fd, temporary_name = tempfile.mkstemp(prefix=f".{target.name}.", suffix=".csv", dir=str(target.parent))
            os.close(fd)
            temporary = Path(temporary_name)
            with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream, delimiter=",").writerows(rows)
            with temporary.open("r", encoding="utf-8-sig", newline="") as stream:
                check = list(csv.reader(stream, delimiter=","))
            if check != rows:
                raise ValueError("CSV_CONTENT_VERIFICATION_FAILED")
            os.replace(str(temporary), str(target))
            return {
                "success": True,
                "text": "CSV успешно создан.",
                "metadata": {
                    "operation": "create_csv",
                    "target_path": str(target),
                    "size_bytes": target.stat().st_size,
                    "row_count": len(rows),
                    "column_count": len(rows[0]),
                    "encoding": "utf-8-sig",
                    "delimiter": ",",
                },
            }
        except Exception as exc:
            if "temporary" in locals() and temporary.exists():
                temporary.unlink()
            return {"success": False, "error": "CREATE_FAILED", "text": f"Не удалось создать CSV: {exc}", "metadata": {}}

    def create_xlsx(self, target, rows):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            return {"success": False, "error": "DEPENDENCY_MISSING", "text": "Библиотека openpyxl недоступна.", "metadata": {}}

        target = Path(str(target))
        if not target.is_absolute() or target.suffix.lower() != ".xlsx":
            return {"success": False, "error": "INVALID_PATH", "text": "Требуется абсолютный Windows-путь к файлу .xlsx.", "metadata": {}}
        if target.exists():
            return {"success": False, "error": "XLSX_TARGET_EXISTS", "text": f"Целевой файл уже существует: {target}", "metadata": {}}
        if not rows or not rows[0] or any(len(row) != len(rows[0]) for row in rows):
            return {"success": False, "error": "CREATE_FAILED", "text": "Табличные данные пусты или имеют разное число столбцов.", "metadata": {}}

        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            fd, temporary_name = tempfile.mkstemp(prefix=f".{target.name}.", suffix=".xlsx", dir=str(target.parent))
            os.close(fd)
            temporary = Path(temporary_name)
            workbook = Workbook()
            sheet = workbook.active
            for row in rows:
                sheet.append(row)
            workbook.save(temporary)
            check = load_workbook(temporary, read_only=True, data_only=False)
            values = [["" if value is None else str(value) for value in row] for row in check.active.iter_rows(values_only=True)]
            check.close()
            if values != rows:
                raise ValueError("XLSX_CONTENT_VERIFICATION_FAILED")
            os.replace(str(temporary), str(target))
            return {
                "success": True,
                "text": "XLSX успешно создан.",
                "metadata": {
                    "operation": "create_xlsx",
                    "target_path": str(target),
                    "size_bytes": target.stat().st_size,
                    "row_count": len(rows),
                    "column_count": len(rows[0]),
                },
            }
        except Exception as exc:
            if "temporary" in locals() and temporary.exists():
                temporary.unlink()
            return {"success": False, "error": "CREATE_FAILED", "text": f"Не удалось создать XLSX: {exc}", "metadata": {}}

    def extract(self, path: Path):

        ext = path.suffix.lower()

        try:

            if ext == ".csv":
                import pandas as pd

                df = pd.read_csv(path)

                return {
                    "success": True,
                    "text": df.to_string(index=False),
                    "metadata": {
                        "handler": "SpreadsheetHandler",
                        "format": "csv",
                        "rows": len(df),
                        "columns": len(df.columns),
                        "headers": list(df.columns)
                    }
                }

            if ext == ".xlsx":
                import pandas as pd

                sheets = pd.read_excel(path, sheet_name=None)

                text_blocks = []

                total_rows = 0

                for name, df in sheets.items():

                    total_rows += len(df)

                    text_blocks.append(f"=== SHEET: {name} ===")
                    text_blocks.append(df.to_string(index=False))

                return {
                    "success": True,
                    "text": "\n\n".join(text_blocks),
                    "metadata": {
                        "handler": "SpreadsheetHandler",
                        "format": "xlsx",
                        "sheet_count": len(sheets),
                        "rows": total_rows,
                        "sheet_names": list(sheets.keys())
                    }
                }

        except Exception as e:

            return {
                "success": False,
                "text": "",
                "metadata": {
                    "handler": "SpreadsheetHandler",
                    "error": str(e)
                }
            }

        return {
            "success": False,
            "text": "",
            "metadata": {
                "handler": "SpreadsheetHandler"
            }
        }
